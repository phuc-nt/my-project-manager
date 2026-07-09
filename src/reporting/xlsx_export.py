"""Build .xlsx report bytes from analyzer dataclasses.

Deterministic and side-effect free: given the already-computed dataclasses
(`ResourceReport`/`CostSummary`, `OkrRollup`) plus an injected `report_date`,
each builder returns `.xlsx` bytes. No network, no clock, no gateway, no LLM —
the numbers come straight from the dataclasses, so the output is reproducible
and unit-testable by re-reading the cells.

`artifact_path` is the single source of truth for where a report file lives on
disk (`data_dir/artifacts/<kind>-<date>.xlsx`); P2's Lớp A confinement check and
P3's write step both derive the location from it.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.agent.okr_analyzer import OkrRollup
from src.tools.models import CostSummary, ResourceReport


def artifact_path(data_dir: Path, kind: str, report_date: str) -> Path:
    """Return the on-disk path for a report artifact, creating the dir.

    `<data_dir>/artifacts/<kind>-<report_date>.xlsx`. The artifacts dir sits
    beside the existing `data_dir/budget` and `data_dir/audit` subdirs. Both the
    attachment-confinement check (P2) and the write step (P3) resolve the file
    location through this one helper so they can never drift apart.
    """
    artifacts_dir = data_dir / "artifacts"
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    return artifacts_dir / f"{kind}-{report_date}.xlsx"


def _workbook_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_resource_xlsx(
    report: ResourceReport,
    cost: CostSummary,
    *,
    report_date: str,
) -> bytes:
    """Resource/cost report → .xlsx bytes.

    One "Resource" sheet: a per-assignee load table (rows preserve the analyzer's
    most-loaded-first order), a spacer, then a labeled cost block. The labor rows
    are omitted when `cost.cost_per_issue == 0` (mirrors the render-layer "n/a"
    rule — labor is not configured).
    """
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Resource"

    ws.append([f"Resource report — {report_date}"])
    ws.append(["Assignee", "Open", "Overdue", "Blockers", "Overloaded"])
    for load in report.loads:
        ws.append(
            [
                load.assignee,
                load.open_count,
                load.overdue_count,
                load.blocker_count,
                "yes" if load.overloaded else "no",
            ]
        )

    ws.append([])
    ws.append(["Cost"])
    ws.append(["Team mean", report.team_mean])
    ws.append(["Unassigned", report.unassigned_count])
    ws.append(["LLM spent", cost.llm_spent])
    ws.append(["LLM cap", cost.llm_cap])
    ws.append(["LLM ratio", cost.llm_ratio])
    ws.append(["LLM status", cost.llm_status])
    ws.append(["Open issues", cost.open_issue_count])
    if cost.cost_per_issue != 0:
        ws.append(["Cost per issue", cost.cost_per_issue])
        ws.append(["Labor estimate", cost.labor_estimate])

    return _workbook_bytes(wb)


def build_okr_xlsx(rollup: OkrRollup, *, report_date: str) -> bytes:
    """OKR rollup → .xlsx bytes.

    One "OKR" sheet, one row per Key Result under its Objective. The objective %
    is repeated on each of its KR rows (KISS — no merged cells). A `None`
    progress renders as an empty cell (distinct from 0). A "Problems" sheet is
    added only when the rollup surfaced skipped rows.
    """
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "OKR"

    ws.append([f"OKR report — {report_date}"])
    ws.append(
        ["Objective", "Key Result", "Epic Keys", "Weight", "KR %", "Objective %"]
    )
    for objective in rollup.objectives:
        for kr in objective.key_results:
            ws.append(
                [
                    objective.name,
                    kr.description,
                    ", ".join(kr.epic_keys),
                    kr.weight,
                    kr.progress_pct,
                    objective.progress_pct,
                ]
            )

    if rollup.problems:
        problems: Worksheet = wb.create_sheet("Problems")
        problems.append(["Row", "Reason"])
        for problem in rollup.problems:
            problems.append([problem.row, problem.reason])

    return _workbook_bytes(wb)
