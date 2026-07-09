"""Unit tests for the xlsx report builders.

Each test builds analyzer dataclasses, calls a builder, then re-opens the
returned bytes with `load_workbook` and asserts exact cell values — proving the
numbers survive the round-trip and match the source dataclass.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from src.agent.okr_analyzer import OkrRollup
from src.reporting.xlsx_export import (
    artifact_path,
    build_okr_xlsx,
    build_resource_xlsx,
)
from src.tools.models import (
    AssigneeLoad,
    CostSummary,
    KeyResult,
    Objective,
    OkrProblem,
    ResourceReport,
)

D = "2026-07-10"


def _resource() -> ResourceReport:
    return ResourceReport(
        loads=(
            AssigneeLoad("Bob", 9, 2, 1, overloaded=True),
            AssigneeLoad("Alice", 3, 0, 0, overloaded=False),
        ),
        team_mean=6.0,
        overloaded=("Bob",),
        unassigned_count=2,
    )


def _sheet_rows(data: bytes, sheet: str) -> list[list]:
    wb = load_workbook(BytesIO(data))
    ws = wb[sheet]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_resource_xlsx_load_table_matches_dataclass():
    cost = CostSummary(12.5, 50.0, 0.25, "ok", 40.0, 8, 5.0)
    rows = _sheet_rows(build_resource_xlsx(_resource(), cost, report_date=D), "Resource")

    assert rows[0] == [f"Resource report — {D}", None, None, None, None]
    assert rows[1] == ["Assignee", "Open", "Overdue", "Blockers", "Overloaded"]
    # Order preserved (most-loaded-first from analyzer).
    assert rows[2] == ["Bob", 9, 2, 1, "yes"]
    assert rows[3] == ["Alice", 3, 0, 0, "no"]


def test_resource_xlsx_cost_block_and_labor_present():
    cost = CostSummary(12.5, 50.0, 0.25, "ok", 40.0, 8, 5.0)
    rows = _sheet_rows(build_resource_xlsx(_resource(), cost, report_date=D), "Resource")
    flat = {r[0]: r[1] for r in rows if r and r[0] and r[1] is not None}

    assert flat["Team mean"] == 6.0
    assert flat["Unassigned"] == 2
    assert flat["LLM spent"] == 12.5
    assert flat["LLM cap"] == 50.0
    assert flat["LLM ratio"] == 0.25
    assert flat["LLM status"] == "ok"
    assert flat["Open issues"] == 8
    assert flat["Cost per issue"] == 5.0
    assert flat["Labor estimate"] == 40.0


def test_resource_xlsx_omits_labor_when_cost_per_issue_zero():
    cost = CostSummary(1.0, 50.0, 0.02, "ok", 0.0, 8, 0.0)
    rows = _sheet_rows(build_resource_xlsx(_resource(), cost, report_date=D), "Resource")
    labels = [r[0] for r in rows if r]

    assert "Cost per issue" not in labels
    assert "Labor estimate" not in labels
    # Rows above the labor block are still present.
    assert "Open issues" in labels


def test_okr_xlsx_one_row_per_kr_objective_pct_repeated():
    rollup = OkrRollup(
        objectives=(
            Objective(
                "Grow retention",
                (
                    KeyResult("Onboarding", ("E-1", "E-2"), 0.7, 60.0),
                    KeyResult("Activation", ("E-3",), 0.3, None),
                ),
                55.0,
            ),
        ),
        problems=(),
        at_risk=(),
    )
    rows = _sheet_rows(build_okr_xlsx(rollup, report_date=D), "OKR")

    assert rows[1] == ["Objective", "Key Result", "Epic Keys", "Weight", "KR %", "Objective %"]
    assert rows[2] == ["Grow retention", "Onboarding", "E-1, E-2", 0.7, 60.0, 55.0]
    # None KR % → empty cell (not 0); objective % repeated on the continuation row.
    assert rows[3] == ["Grow retention", "Activation", "E-3", 0.3, None, 55.0]


def test_okr_xlsx_problems_sheet_only_when_present():
    clean = OkrRollup(
        objectives=(Objective("O", (KeyResult("KR", ("E-1",), None, 10.0),), 10.0),),
        problems=(),
        at_risk=(),
    )
    wb = load_workbook(BytesIO(build_okr_xlsx(clean, report_date=D)))
    assert wb.sheetnames == ["OKR"]

    with_problem = OkrRollup(
        objectives=(Objective("O", (KeyResult("KR", ("E-1",), None, 10.0),), 10.0),),
        problems=(OkrProblem("O | KR", "epic E-9 not found"),),
        at_risk=(),
    )
    prob_rows = _sheet_rows(build_okr_xlsx(with_problem, report_date=D), "Problems")
    assert prob_rows[0] == ["Row", "Reason"]
    assert prob_rows[1] == ["O | KR", "epic E-9 not found"]


def test_artifact_path_shape_and_dir_creation(tmp_path: Path):
    p = artifact_path(tmp_path, "resource", D)
    assert p == tmp_path / "artifacts" / f"resource-{D}.xlsx"
    assert p.parent.is_dir()  # created eagerly
